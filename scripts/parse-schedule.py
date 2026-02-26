#!/usr/bin/env python3
"""
parse-schedule.py
Reads complete-table.xlsx and produces schedule.json ready for database seeding.

The complete-table.xlsx uses SHORTHAND NICKNAMES (e.g., กล้วย, ท็อป, ฟอร์ม, มุก).
All names in the table are pharmacists — no filtering needed since this is
the pharmacist-only schedule table.

Usage:
  cd /path/to/pharmshift
  python3 scripts/parse-schedule.py

Output:
  scripts/schedule.json
"""

import json
import os
import sys
import re
import datetime
from pathlib import Path

try:
    import openpyxl
    import xlrd
except ImportError:
    print("Installing required packages...")
    os.system("pip3 install openpyxl xlrd")
    import openpyxl
    import xlrd

# ── Paths ────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
ROOT_DIR     = SCRIPT_DIR.parent
EXCEL_DIR    = ROOT_DIR / "Excel"
COMPLETE_XLS = EXCEL_DIR / "complete-table.xlsx"
OUTPUT_JSON  = SCRIPT_DIR / "schedule.json"

# ── Keywords to skip (not names) ─────────────────────────────
NON_NAME_TOKENS = {
    'อาทิตย์','จันทร์','อังคาร','พุธ','พฤหัสบดี','ศุกร์','เสาร์',
    'โครงการ','SURG','MED','ER','SMC','รุ่งอรุณ',
    'บ่าย','ดึก', 'ANA',
}
TITLE_PREFIX = 'ตารางเวร'

# Column layouts (1-indexed) — day of week column starts
DAY_COL_STARTS = [1, 6, 10, 14, 18, 22, 26]  # Sun→Sat

# ── Helpers ──────────────────────────────────────────────────
def cell_val(ws, row, col):
    """Safe cell value read."""
    try:
        v = ws.cell(row, col).value
        if isinstance(v, str):
            return v.strip().replace('\n', '').replace('\r', '').replace(' ', '')
        return v
    except:
        return None

def is_dept(v):
    return isinstance(v, str) and v in {'โครงการ','SURG','MED','ER','SMC','รุ่งอรุณ'}

def is_shift(v):
    return isinstance(v, str) and v in {'บ่าย','ดึก'}

def is_day_num(v):
    return isinstance(v, (int, float)) and 1 <= v <= 31

def parse_name(v):
    """Clean a name cell; return None if it's a keyword or empty."""
    if not isinstance(v, str):
        return None
    cleaned = v.strip().replace('\n','').replace('\r','').replace('  ',' ')
    if not cleaned:
        return None
    if cleaned in NON_NAME_TOKENS:
        return None
    if cleaned.startswith(TITLE_PREFIX):
        return None
    # Skip if it looks like a number
    try:
        float(cleaned)
        return None
    except:
        pass
    return cleaned

# ── Main parse logic ─────────────────────────────────────────
def parse_complete_table():
    wb = openpyxl.load_workbook(str(COMPLETE_XLS), data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # --- Detect month/year from row 1 ---
    title = ws.cell(1, 1).value or ''
    month_map = {
        'มกราคม':1,'กุมภาพันธ์':2,'มีนาคม':3,'เมษายน':4,
        'พฤษภาคม':5,'มิถุนายน':6,'กรกฎาคม':7,'สิงหาคม':8,
        'กันยายน':9,'ตุลาคม':10,'พฤศจิกายน':11,'ธันวาคม':12,
    }
    thai_month = 3
    thai_year  = 2569
    for mname, mnum in month_map.items():
        if mname in title:
            thai_month = mnum
            break
    for tok in str(title).split():
        if tok.isdigit() and len(tok) == 4:
            thai_year = int(tok)
    ce_year  = thai_year - 543
    ce_month = thai_month
    print(f'Schedule: {thai_month}/{thai_year} → CE {ce_month}/{ce_year}')

    shifts = []
    all_names = set()

    # Scan whole table row by row, building context per day column
    # Strategy: scan each row, look at each day column block (5 cols wide)

    row_data = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, 36):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                v = v.strip().replace('\n','').replace('\r','')
            row.append(v)
        row_data.append(row)

    # Find week-block header rows (rows where col 4 = integer day number)
    # These are the rows like: [โครงการ, SURG, MED, บ่าย, 1, โครงการ, SMC, บ่าย, 2, ...]
    def is_header_row(row):
        # Heuristic: several day number integers in specific positions
        count = 0
        for ci in [4, 8, 12, 16, 20, 24, 29]:  # 0-indexed
            v = row[ci] if ci < len(row) else None
            if is_day_num(v):
                count += 1
        return count >= 3

    header_rows = []  # 0-indexed
    for ri, row in enumerate(row_data):
        if is_header_row(row):
            header_rows.append(ri)

    print(f'Header rows (0-idx): {header_rows}')

    for block_idx, hrow_idx in enumerate(header_rows):
        hrow = row_data[hrow_idx]

        # This block spans from hrow_idx to next header_row - 1 (or end)
        next_hrow = header_rows[block_idx + 1] if block_idx + 1 < len(header_rows) else len(row_data)
        block = row_data[hrow_idx : next_hrow]  # list of rows

        # Process each of the 7 day columns
        for day_idx, col_start in enumerate(DAY_COL_STARTS):
            ci = col_start - 1  # 0-indexed

            # Get day number from header row (5th cell of the day block = index ci+4)
            day_num = None
            for extra in range(4, -1, -1):
                v = hrow[ci + extra] if (ci + extra) < len(hrow) else None
                if is_day_num(v):
                    day_num = int(v)
                    break

            if day_num is None:
                continue

            try:
                the_date = datetime.date(ce_year, ce_month, day_num)
            except ValueError:
                continue

            date_str = the_date.isoformat()

            # Determine สด dept (บ่าย) from header row cells
            bai_dept = 'โครงการ'  # default
            for extra in range(4):
                v = hrow[ci + extra] if (ci + extra) < len(hrow) else None
                if is_dept(v):
                    bai_dept = v
                    break

            # Now scan subsequent rows in the block for names and for ดึก section
            bai_names  = []
            duek_names = []
            duek_dept  = 'ER'  # default
            in_duek    = False

            for row_offset in range(1, len(block)):
                brow = block[row_offset]
                row_main = brow[ci] if ci < len(brow) else None

                # Check if this sub-row starts a ดึก section for this day
                # The ดึก header appears as: ER/รุ่งอรุณ ดึก ... in the window
                window = [brow[ci + e] if (ci + e) < len(brow) else None for e in range(4)]
                has_duek = 'ดึก' in window
                has_dept = any(is_dept(v) for v in window)

                if has_duek and has_dept:
                    in_duek = True
                    for v in window:
                        if is_dept(v):
                            duek_dept = v
                            break
                    continue

                # Collect names
                name = parse_name(row_main)
                if name:
                    all_names.add(name)
                    if in_duek:
                        duek_names.append(name)
                    else:
                        bai_names.append(name)

            # Build shift rows
            for name in bai_names:
                shifts.append({
                    'date': date_str,
                    'nickname': name,
                    'shift_type': 'บ่าย',
                    'department': bai_dept,
                    'month_year': f'{ce_year}-{str(ce_month).zfill(2)}',
                })
            for name in duek_names:
                shifts.append({
                    'date': date_str,
                    'nickname': name,
                    'shift_type': 'ดึก',
                    'department': duek_dept,
                    'month_year': f'{ce_year}-{str(ce_month).zfill(2)}',
                })

    print(f'Total shifts: {len(shifts)}')
    print(f'Unique nicknames: {sorted(all_names)}')
    return shifts, all_names, f'{ce_year}-{str(ce_month).zfill(2)}'

def main():
    print('Parsing schedule...')
    shifts, all_names, month_year = parse_complete_table()

    # Build user list from all unique nicknames
    users = [
        {
            'nickname': n,
            'name': n,       # Will be overridden when linked to full-name list
            'prefix': '',
            'profile_image': 'male',
        }
        for n in sorted(all_names)
    ]

    bai_count  = sum(1 for s in shifts if s['shift_type'] == 'บ่าย')
    duek_count = sum(1 for s in shifts if s['shift_type'] == 'ดึก')

    output = {
        'month_year': month_year,
        'users': users,
        'shifts': shifts,
        'summary': {
            'total_shifts': len(shifts),
            'total_users': len(users),
            'bai_shifts': bai_count,
            'duek_shifts': duek_count,
        }
    }

    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f'\n✅ Saved → {OUTPUT_JSON}')
    print(f'   Users (pharmacist nicknames): {len(users)}')
    print(f'   Shifts: {len(shifts)} (บ่าย: {bai_count}, ดึก: {duek_count})')

if __name__ == '__main__':
    main()
